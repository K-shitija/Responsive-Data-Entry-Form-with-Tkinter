# -*- coding: utf-8 -*-
"""
Created on Thu Feb  1 09:30:33 2024

@author: KSHITIJA
"""

import tkinter 
from tkinter import ttk
#ttk = themed tkinter
from tkinter import messagebox
import os
import openpyxl
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side,NamedStyle
from openpyxl.utils import get_column_letter


def clear_entry():
    first_name_entry.delete(0,tkinter.END)
    
    last_name_entry.delete(0,tkinter.END)
    title_combobox.delete(0,tkinter.END)
    age_spinbox.delete(0,tkinter.END)
    nationality_combobox.delete(0,tkinter.END)
    #reg_status_var.delete(0,tkinter.END)
    reg_status_var = tkinter.StringVar(value="Not Registered")
    numcourses_spinbox.delete(0,tkinter.END)
    numsemesters_spinbox.delete(0,tkinter.END)
    
    
    
    
    
def enter_data():
    accepted = accept_var.get()
    
    if accepted=="Accepted":
        # User info
        firstname = first_name_entry.get()
        lastname = last_name_entry.get()
        
        if firstname and lastname:
            title = title_combobox.get()
            age = age_spinbox.get()
            nationality = nationality_combobox.get()
            
            # Course info
            registration_status = reg_status_var.get()
            numcourses = numcourses_spinbox.get()
            numsemesters = numsemesters_spinbox.get()
            
            print("First name: ", firstname, "Last name: ", lastname)
            print("Title: ", title, "Age: ", age, "Nationality: ", nationality)
            print("# Courses: ", numcourses, "# Semesters: ", numsemesters)
            print("Registration status", registration_status)
            print("------------------------------------------")
            
            filepath = "F:\Python_project_persoanl\Registration form\Registration_data.xlsx"
            
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Registration_data"
                heading = ["First Name", "Last Name", "Title", "Age", "Nationality",
                           "# Courses", "# Semesters", "Registration status"]
                sheet.cell(row=1,column=1).value = "First Name"
                sheet.cell(row=1,column=2).value = "Last Name"
                sheet.cell(row=1,column=3).value = "Title "
                sheet.cell(row=1,column=4).value = "Age"
                sheet.cell(row=1,column=5).value = "Nationality"
                sheet.cell(row=1,column=6).value = "# Courses"
                sheet.cell(row=1,column=7).value = "# Semesters"
                sheet.cell(row=1,column=8).value = "Registration status"
                
                #sheet.append(heading)
                for i, header in enumerate(heading, start=1):
                    a_cell= sheet.cell(row=1, column=i)
                    
                    font = Font(name = "Albany AMT",size=9.5, bold=True, color= '00008B')
                    
                    a_cell.font = font
                    
                    fill = PatternFill(start_color = 'FFd1e7f0', end_color = 'FFd1e7f0', fill_type = 'solid')
                    a_cell.fill = fill
                    a_cell.border = Border(left = Side(style='thin'),right = Side(style='thin'),top= Side(style='thin')
                                           ,bottom= Side(style='thin'))
                    sheet.column_dimensions[get_column_letter(i)].width =20
                
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([firstname, lastname, title, age, nationality, numcourses,
                          numsemesters, registration_status])
            workbook.save(filepath)
            
                    
                   
        
        else:
            tkinter.messagebox.showwarning(title="Error", message="First name and last name are required.")
    else:
        tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms")
    clear_entry()
            
            
            
            

window = tkinter.Tk()
window.title("Data Entry Form")

frame = tkinter.Frame(window)
frame.pack()

# Saving User Info
user_info_frame =tkinter.LabelFrame(frame, text="User Information")
user_info_frame.grid(row= 0, column=0, padx=20, pady=10)

first_name_label = tkinter.Label(user_info_frame, text="First Name")
first_name_label.grid(row=0, column=0)
last_name_label = tkinter.Label(user_info_frame, text="Last Name")
last_name_label.grid(row=0, column=1)

first_name_entry = tkinter.Entry(user_info_frame)
last_name_entry = tkinter.Entry(user_info_frame)
first_name_entry.grid(row=1,column=0)
last_name_entry.grid(row=1,column=1)

title = tkinter.Label(user_info_frame , text ="Title")
#we can add here font, color ,size ,bg
title.grid(row=0, column =2)
title_combobox= ttk.Combobox(user_info_frame, values =["Mr", "Ms","Mrs","Dr","Major","LT", ""])
title_combobox.grid(row=1,column=2)

age_label = tkinter.Label(user_info_frame, text="Age")
age_spinbox = tkinter.Spinbox(user_info_frame, from_=18, to=110)
age_label.grid(row=2, column=0)
age_spinbox.grid(row=3, column=0)

nationality_label = tkinter.Label(user_info_frame, text="Nationality")
nationality_combobox = ttk.Combobox(user_info_frame, values=["Africa", "Antarctica", "Asia", "Europe", "North America", "Oceania", "South America"])
nationality_label.grid(row=2, column=1)
nationality_combobox.grid(row=3, column=1)


# for proper spacing 
for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)
    
# Saving Course Info
courses_frame = tkinter.LabelFrame(frame)
courses_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)
 #sticky : to expand the course frame inline with user-info -frame , news=north ,east,west,south
registered_label = tkinter.Label(courses_frame, text="Registration Status")

reg_status_var = tkinter.StringVar(value="Not Registered")
registered_check = tkinter.Checkbutton(courses_frame, text="Currently Registered",
                                       variable=reg_status_var, onvalue="Registered", offvalue="Not registered")

registered_label.grid(row=0, column=0)
registered_check.grid(row=1, column=0)

numcourses_label = tkinter.Label(courses_frame, text= "# Completed Courses")
numcourses_spinbox = tkinter.Spinbox(courses_frame, from_=0, to='infinity')
numcourses_label.grid(row=0, column=1)
numcourses_spinbox.grid(row=1, column=1)

numsemesters_label = tkinter.Label(courses_frame, text="# Semesters")
numsemesters_spinbox = tkinter.Spinbox(courses_frame, from_=0, to="infinity")
numsemesters_label.grid(row=0, column=2)
numsemesters_spinbox.grid(row=1, column=2)

for widget in courses_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)
    
    # Accept terms

terms_frame = tkinter.LabelFrame(frame, text="Terms & Conditions")
terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

accept_var = tkinter.StringVar(value="Not Accepted")
terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
                                  variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=0, column=0)

# Button
button = tkinter.Button(frame, text="Enter data", command= enter_data)
button.grid(row=3, column=0, sticky="news", padx=20, pady=10)



window.mainloop()
